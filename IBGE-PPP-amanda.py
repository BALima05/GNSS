import os
import zipfile
import shutil
from datetime import datetime
from openpyxl import Workbook
from tkinter import Tk, filedialog

def obter_data_inicio(arquivo_sum):
    with open(arquivo_sum, 'r', encoding='utf-8', errors='ignore') as file:
        for linha in file:
            if linha.strip().startswith("Inicio"):
                partes = linha.strip().split(":")
                if len(partes) > 1:
                    data_str = partes[1].strip().split(" ")[0]  # pega apenas AAAA/MM/DD
                    try:
                        data_obj = datetime.strptime(data_str, "%Y/%m/%d")
                        return data_obj.strftime("%d/%m/%Y")
                    except ValueError:
                        pass
    return ""

def formatar_gms(valor):
    partes = valor.split()
    if len(partes) == 3:
        return f"{partes[0]}° {partes[1]}' {partes[2]}\""
    return valor

def extrair_informacoes_text(arquivo_text):
    informacoes = {}
    with open(arquivo_text, 'r', encoding='utf-8', errors='ignore') as file:
        for linha in file:
            partes = linha.split()
            if linha.startswith("SLAT"):
                informacoes["Sigma Latitude (95%) (m)"] = partes[1]
            elif linha.startswith("SLON"):
                informacoes["Sigma Longitude (95%) (m)"] = partes[1]
            elif linha.startswith("SHGEO"):
                informacoes["Sigma Alt. Geo. (95%) (m)"] = partes[1]
    return informacoes

def extrair_informacoes_sum(arquivo_sum):
    informacoes = {
        "Resíduos da pseudo distância GPS (m)": None,
        "Resíduos da pseudo distância GLONASS (m)": None,
        "Resíduos da fase da portadora GPS (cm)": None,
        "Resíduos da fase da portadora GLONASS (cm)": None,
        "Latitude(gms)": None,
        "Longitude(gms)": None,
        "Altitude Geométrica (m)": None
    }
    with open(arquivo_sum, 'r', encoding='utf-8', errors='ignore') as file:
        for linha in file:
            partes = linha.split()
            if "Residuos da pseudodistancia" in linha:
                valor = partes[-2].replace(',', '.')
                if "GPS" in linha:
                    informacoes["Resíduos da pseudo distância GPS (m)"] = valor
                elif "GLONASS" in linha:
                    informacoes["Resíduos da pseudo distância GLONASS (m)"] = valor
            elif "Residuos da fase da portadora" in linha:
                valor = partes[-2].replace(',', '.')
                if "GPS" in linha:
                    informacoes["Resíduos da fase da portadora GPS (cm)"] = valor
                elif "GLONASS" in linha:
                    informacoes["Resíduos da fase da portadora GLONASS (cm)"] = valor
            elif "Latitude  (gms)" in linha:
                informacoes["Latitude(gms)"] = formatar_gms(" ".join(partes[2:5]))
            elif "Longitude (gms)" in linha:
                informacoes["Longitude(gms)"] = formatar_gms(" ".join(partes[2:5]))
            elif "Alt. Geo. (m)" in linha and len(partes) >= 3:
                try:
                    for valor in partes:
                        try:
                            informacoes["Altitude Geométrica (m)"] = float(valor.replace(',', '.'))
                            break
                        except ValueError:
                            continue
                except Exception:
                    continue
    return informacoes

def descompactar_sem_subpastas():
    Tk().withdraw()
    arquivos_zip = filedialog.askopenfilenames(
        title="Selecione os arquivos ZIP",
        filetypes=[("ZIP files", "*.zip")]
    )

    if not arquivos_zip:
        print("Nenhum arquivo selecionado.")
        return None

    pasta_saida = os.path.join(os.path.dirname(arquivos_zip[0]), "Arquivos descompactados")
    os.makedirs(pasta_saida, exist_ok=True)

    for zip_path in arquivos_zip:
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                for membro in zip_ref.infolist():
                    if membro.is_dir():
                        continue
                    nome_arquivo = os.path.basename(membro.filename)
                    if not nome_arquivo:
                        continue
                    caminho_destino = os.path.join(pasta_saida, nome_arquivo)
                    base, ext = os.path.splitext(nome_arquivo)
                    contador = 1
                    while os.path.exists(caminho_destino):
                        nome_arquivo = f"{base}_{contador}{ext}"
                        caminho_destino = os.path.join(pasta_saida, nome_arquivo)
                        contador += 1
                    with zip_ref.open(membro) as fonte, open(caminho_destino, 'wb') as destino:
                        shutil.copyfileobj(fonte, destino)
                print(f"Descompactado: {os.path.basename(zip_path)}")
        except zipfile.BadZipFile:
            print(f"Arquivo corrompido ou inválido: {os.path.basename(zip_path)}")

    return pasta_saida

def procurar_informacoes(pasta_arquivos):
    dados_gps = []
    dados_glonass = []
    dados_duplo = []

    for arquivo in os.listdir(pasta_arquivos):
        if arquivo.endswith(".sum"):
            caminho_sum = os.path.join(pasta_arquivos, arquivo)
            print(f"Analisando: {arquivo}")

            nome_arquivo = "poli" + arquivo.split("poli")[1].split(".")[0]
            tipo = (
                "GPS E GLONASS" if ("GPS" in arquivo.upper() and "GLONASS" in arquivo.upper())
                else "GPS" if "GPS" in arquivo.upper()
                else "GLONASS"
            )

            arquivo_text = arquivo.replace(".sum", ".txt")
            caminho_text = os.path.join(pasta_arquivos, arquivo_text)

            if not os.path.exists(caminho_text):
                print(f"Arquivo correspondente {arquivo_text} não encontrado. Pulando...")
                continue

            info_text = extrair_informacoes_text(caminho_text)
            info_sum = extrair_informacoes_sum(caminho_sum)
            data = obter_data_inicio(caminho_sum)

            dados = {
                "Nome do arquivo": nome_arquivo,
                "Tipo": tipo,
                "Data": data,
                **info_text,
                **info_sum
            }

            if tipo == "GPS":
                dados_gps.append(dados)
            elif tipo == "GLONASS":
                dados_glonass.append(dados)
            else:
                dados_duplo.append(dados)

    todas_colunas = [
        "Nome do arquivo", "Tipo", "Data",
        "Latitude(gms)", "Sigma Latitude (95%) (m)",
        "Longitude(gms)", "Sigma Longitude (95%) (m)",
        "Sigma Alt. Geo. (95%) (m)", "Altitude Geométrica (m)",
        "Resíduos da pseudo distância GPS (m)",
        "Resíduos da pseudo distância GLONASS (m)",
        "Resíduos da fase da portadora GPS (cm)",
        "Resíduos da fase da portadora GLONASS (cm)"
    ]

    wb = Workbook()
    ws_gps = wb.active
    ws_gps.title = "GPS"
    ws_glonass = wb.create_sheet("GLONASS")
    ws_duplo = wb.create_sheet("GPS E GLONASS")

    def preencher_aba(sheet, dados, tipo):
        if tipo == "GPS":
            colunas = [col for col in todas_colunas if "GLONASS" not in col]
        elif tipo == "GLONASS":
            colunas = [col for col in todas_colunas if "GPS" not in col or col == "Tipo"]
        else:
            colunas = todas_colunas
        sheet.append(colunas)
        for linha in dados:
            row = [linha.get(col, "") for col in colunas]
            sheet.append(row)

    preencher_aba(ws_gps, dados_gps, "GPS")
    preencher_aba(ws_glonass, dados_glonass, "GLONASS")
    preencher_aba(ws_duplo, dados_duplo, "GPS E GLONASS")

    caminho_excel = os.path.join(pasta_arquivos, "informacoes_geograficas.xlsx")
    wb.save(caminho_excel)
    print(f"\nArquivo Excel salvo em: {caminho_excel}")

if __name__ == "__main__":
    pasta = descompactar_sem_subpastas()
    if pasta:
        procurar_informacoes(pasta)
